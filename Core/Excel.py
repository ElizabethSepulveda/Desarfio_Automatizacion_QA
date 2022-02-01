# file: Core/Excel.py
# ----------------------------------------------------------------------------
# MODULE: Interactúa con el archivo Excel (Lee e Inserta)
# ----------------------------------------------------------------------------

import pandas
from openpyxl import load_workbook


# retorna un pandas dataframe con los datos de la hoja del archivo excel recibido
def getDataframe(excel_path, excel_sheet=None):
    if excel_sheet is None:
        excel_sheet = 0
    try:
        return pandas.read_excel(excel_path, sheet_name=excel_sheet, engine="openpyxl").applymap(str)
    except Exception as e:
        print(e)
        raise Exception(e)


# convierte los datos de la hoja del archivo excel recibido en una tabla para el feature
def getTable(behave, excel_path, excel_sheet=None):
    dataframe = getDataframe(excel_path, excel_sheet)
    headings = dataframe.columns.values.tolist()
    examples_table = behave.model.Table(headings=headings)
    for i in dataframe.index:
        cells = dataframe.iloc[[i]].values.tolist()[0]
        row = behave.model.Row(headings=headings, cells=cells)
        examples_table.add_row(row)
    return examples_table


# inserta una celda en una hoja, columna y fila especificados
# si la hoja es None, tomará la primera por defecto
# si la columan no existe, se creará al final de la tabla ya existente
def insertCell(excel_path, data, sheet=None, column_name="data", row=0):
    # apertura del excel
    df = pandas.DataFrame([data], columns=[column_name])
    book = load_workbook(excel_path)
    writer = pandas.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    if sheet is None:
        sheet = book.worksheets[0].title

    # obtención del índice de la columna, si no existe se crea la columna
    # al lado de la ultima columna
    worksheet = book[sheet]
    column_index = -1
    for col in worksheet.iter_cols(1, worksheet.max_column):
        if col[0].value == column_name:
            column_index = col[0].col_idx - 1
            break
    if column_index == -1:
        column_index = worksheet.max_column

    df.to_excel(writer, sheet, startcol=column_index, startrow=row, index=False)
    writer.save()

